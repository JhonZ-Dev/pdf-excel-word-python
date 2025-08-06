import openpyxl
import json
from copy import copy

def save_merges(ws):
    merges = []
    print("\n--- Lista de merges después de restaurar ---")
    for cr in ws.merged_cells.ranges:
        print(str(cr))
        merges.append({
            'min_row': cr.min_row,
            'max_row': cr.max_row,
            'min_col': cr.min_col,
            'max_col': cr.max_col,
        })
    return merges

def restore_merges(ws, merges, insert_at, n_rows, item_start_row, new_total_row, max_col):
    """
    Restaura merges, ajustando filas si el merge está después de insert_at.
    Evita merges que toquen cualquier fila de la tabla de items y total.
    """
    for m in merges:
        min_row = m['min_row']
        max_row = m['max_row']
        min_col = m['min_col']
        max_col_merge = m['max_col']
        # Si el merge está después de donde insertaste filas, desplázalo
        if min_row >= insert_at:
            min_row += n_rows
            max_row += n_rows
        # NO recrees el merge en la fila de TOTAL original (solo lo ponemos en la nueva fila TOTAL después)
        if insert_at <= m['min_row'] <= m['max_row'] <= insert_at:
            continue
        # 🚨 SKIP merges que toquen cualquier parte de la tabla de items o la fila total (A14:G[new_total_row])
        if (
            max_row >= item_start_row and   # Toca o pasa sobre inicio de items
            min_row <= new_total_row and    # Toca o pasa sobre fin de items
            min_col <= max_col and
            max_col_merge >= 1
        ):
            continue  # ¡Este merge toca la tabla de items, lo saltamos!
        try:
            ws.merge_cells(
                start_row=min_row, start_column=min_col,
                end_row=max_row, end_column=max_col_merge
            )
        except:
            pass



with open('datos.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = openpyxl.load_workbook('template.xlsx')
ws = wb.active

cabecera_map = {
    'txt_cliente':     'D2',
    'txt_ruc':         'D3',
    'txt_direccion':   'D4',
    'txt_fecha':       'D5',
    'txt_telefono':    'D6',
    'txt_necesidad':   'D7',
    'txt_funcionario': 'D8',
    'txt_correo':      'D9',
    'tHora_maxina':    'D10',
}
for k, cell in cabecera_map.items():
    ws[cell] = data.get(k, "")

n_items = len(data['items'])
item_start_row = 14
total_row_original = 15
max_col = 7

# 1. Guarda todos los merges originales (con coordenadas, NO como string)
merges_orig = save_merges(ws)

# 2. Insertar filas antes de TOTAL (si hay más de 2 items)
rows_to_insert = max(0, n_items - 2)
if rows_to_insert:
    ws.insert_rows(total_row_original, rows_to_insert)

new_total_row = item_start_row + n_items

# 3. Elimina TODOS los merges (temporalmente)
for rng in list(ws.merged_cells.ranges):
    try:
        ws.unmerge_cells(str(rng))
    except:
        pass

# 4. Restaura merges desplazando los que están después de la fila TOTAL original
#restore_merges(ws, merges_orig, insert_at=total_row_original, n_rows=rows_to_insert)
restore_merges(
    ws, merges_orig,
    insert_at=total_row_original,
    n_rows=rows_to_insert,
    item_start_row=item_start_row,
    new_total_row=new_total_row,
    max_col=max_col
)
to_unmerge = []
for cr in list(ws.merged_cells.ranges):
    if (
        cr.max_row >= item_start_row and
        cr.min_row <= new_total_row and
        cr.max_col >= 1 and
        cr.min_col <= max_col
    ):
        to_unmerge.append(str(cr))

for ref in to_unmerge:
    try:
        ws.unmerge_cells(ref)
    except:
        pass
# 5. Copia formato SOLO de la fila 14 original para nuevas filas de items
for i in range(n_items):
    src_row = item_start_row
    dst_row = item_start_row + i
    if i > 0:
        for col in range(1, max_col+1):
            src_cell = ws.cell(row=src_row, column=col)
            dst_cell = ws.cell(row=dst_row, column=col)
            if not isinstance(dst_cell, openpyxl.cell.cell.MergedCell):
                dst_cell._style = copy(src_cell._style)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
for idx in range(n_items):
    row = item_start_row + idx
    print(f"Fila {row} columna D es MergedCell? ", isinstance(ws.cell(row=row, column=4), openpyxl.cell.cell.MergedCell))

# 6. Escribe los items solo si la celda NO es MergedCell
for idx, item in enumerate(data['items']):
    row = item_start_row + idx
    if not isinstance(ws.cell(row=row, column=1), openpyxl.cell.cell.MergedCell):
        ws[f'A{row}'] = idx + 1
    if not isinstance(ws.cell(row=row, column=2), openpyxl.cell.cell.MergedCell):
        ws[f'B{row}'] = item['txt_cpc']
    if not isinstance(ws.cell(row=row, column=3), openpyxl.cell.cell.MergedCell):
        ws[f'C{row}'] = item['txt_unidad']
    if not isinstance(ws.cell(row=row, column=4), openpyxl.cell.cell.MergedCell):
        ws[f'D{row}'] = item['txt_especificaciones']
    if not isinstance(ws.cell(row=row, column=5), openpyxl.cell.cell.MergedCell):
        ws[f'E{row}'] = item['int_cantidad']
    if not isinstance(ws.cell(row=row, column=6), openpyxl.cell.cell.MergedCell):
        ws[f'F{row}'] = item['flo_precioUnitario']
    if not isinstance(ws.cell(row=row, column=7), openpyxl.cell.cell.MergedCell):
        ws[f'G{row}'] = item['flo_precioTotal']

# 7. Vuelve a combinar la fila TOTAL en la nueva ubicación (A[new_total_row]:F[new_total_row])
ws.merge_cells(start_row=new_total_row, start_column=1, end_row=new_total_row, end_column=6)

# Escribir el texto y el total en la fila TOTAL
if not isinstance(ws.cell(row=new_total_row, column=1), openpyxl.cell.cell.MergedCell):
    ws[f'A{new_total_row}'] = "TOTAL"
if not isinstance(ws.cell(row=new_total_row, column=7), openpyxl.cell.cell.MergedCell):
    ws[f'G{new_total_row}'] = sum(float(item['flo_precioTotal']) for item in data['items'])

# Copiar formato de TOTAL original a la nueva fila TOTAL (opcional)
for col in range(1, max_col+1):
    src_cell = ws.cell(row=total_row_original, column=col)
    dst_cell = ws.cell(row=new_total_row, column=col)
    if not isinstance(dst_cell, openpyxl.cell.cell.MergedCell):
        dst_cell._style = copy(src_cell._style)
        dst_cell.number_format = src_cell.number_format
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)

wb.save('proforma_final.xlsx')
